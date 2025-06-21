import openpyxl
import FreeSimpleGUI as sg
import os
import csv

dossier = os.path.dirname(os.path.abspath(__file__))

fichier_licencies = os.path.join(dossier, "licencies.csv")
template_feuille = os.path.join(dossier, "feuille_match_vide.xlsx")

# Coordonnees cellule nom equipe locale
coord_locaux_visiteurs = [[5, 5], [28, 5]]
# premiere ligne des licences en locaux/adversaires
prem_ligne_licences = [15, 38]
# Colonne du nom du joueur
col_joueur = 5
# Colonne du numéro du joueur
col_num = 10

dic_licencies = {}
with open(fichier_licencies, newline='') as csvfile:
    licencies = csv.reader(csvfile, delimiter=';', quotechar='|')
    for l in licencies:
        # On verifie que la ligne n'est pas vide
        if len(l) > 0:
            dic_licencies[l[0]] = l[1:]


def GUI(dic_licencies):
    liste_licencies = list(dic_licencies.keys())
    layout = [
        [sg.T("")],
        [sg.Text("Equipe : "), sg.Input("Drime Time", key="Equipe")],
        [sg.Text("Adversaire : "), sg.Input(key="Adversaire")],
        [sg.Radio('Locaux', group_id=1, default=True, key="Locaux"),
         sg.Radio('Visiteurs', group_id=1)],
        [sg.Text("Sélectionner joueurs"), sg.Listbox(values=liste_licencies, select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE,
                                                     key="Joueurs", enable_events=True, size=(10, len(liste_licencies)))],
        [sg.Button('OK')]]
    window = sg.Window('Paramètrage feuille', layout)
    reglages = {}
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        if event == "OK":
            reglages["Equipe"] = values["Equipe"]
            reglages["Adversaire"] = values["Adversaire"]
            reglages["Locaux"] = values["Locaux"]
            # On ne prend que les joueurs selectionnes parmis les licencies
            reglages["Joueurs"] = {k: dic_licencies[k]
                                   for k in values["Joueurs"]}
            break
    window.close()
    return reglages


def ecrire_feuille(template_feuille, prem_ligne_licences, col_joueur, col_num, reglages):
    [Locaux, Adversaire, Joueurs, Equipe] = [reglages["Locaux"],
                                             reglages["Adversaire"], reglages["Joueurs"], reglages["Equipe"]]
    xlsx = openpyxl.load_workbook(template_feuille, data_only=True)
    feuille = xlsx.active
    # 0 pour Locaux, 1 pour Visiteurs
    index_locaux = int(not Locaux)
    # Ecriture des equipes
    feuille.cell(coord_locaux_visiteurs[index_locaux][0],
                 coord_locaux_visiteurs[index_locaux][1], Equipe)
    feuille.cell(coord_locaux_visiteurs[index_locaux-1][0],
                 coord_locaux_visiteurs[index_locaux-1][1], Adversaire)
    # Ecriture des joueurs
    ligne_joueurs = prem_ligne_licences[index_locaux]
    for v in Joueurs.keys():
        feuille.cell(ligne_joueurs, 1, Joueurs[v][1])
        feuille.cell(ligne_joueurs, col_joueur, Joueurs[v][0])
        feuille.cell(ligne_joueurs, col_num, Joueurs[v][2])
        ligne_joueurs += 1
    xlsx.save(os.path.join(dossier, f"Feuille_{Equipe}_{Adversaire}_.xlsx"))


reglages = GUI(dic_licencies)
ecrire_feuille(template_feuille, prem_ligne_licences,
               col_joueur, col_num, reglages)
